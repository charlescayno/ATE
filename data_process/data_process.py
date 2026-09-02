import math
import pandas as pd
from copy import copy, deepcopy

import openpyxl
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.styles import Alignment
from openpyxl.chart._chart import ChartBase
from openpyxl.chart.data_source import StrRef,AxDataSource, NumRef
from openpyxl.chart.reference import Reference
from openpyxl.chart.series import Series
from openpyxl.chart.series_factory import SeriesFactory
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.colors import ColorChoice
from openpyxl.drawing.text import  Paragraph, ParagraphProperties, CharacterProperties, Font
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils import get_column_letter


from misc_functions.misc_functions import timeit

import os
import shutil
import matplotlib.pyplot as plt

from user_settings.save_load import default_config_folder_path


class ExcelPlotterError(Exception):
    pass

chart_template_path = f"{default_config_folder_path}/ATE_Charts_Template.xlsx"

#create color list for charts with more than 4 series
color_list = ['C00000','FFC000','ED7D31', 'FF3399',
              'FFFF00','FF0000','00EAFF', 'AA00FF',
              'FF7F00','BFFF00','0095FF', 'FF00AA',
              '6AFF00','0040FF','EDB9B9', 'B9D7ED',
              'E7E9B9','DCB9ED','B9EDE0', '8F2323',
              '23628F','8F6A23','6B238F', '4F8F23',
              '737373','CCCCCC','FFFFFF', '000000']

def path_maker(file_path: str):
    """
        file_path: Enter the file path you want to create

        returns the string value of new path created
    """

    folder_list = file_path.split('/')

    new_path = ' '
    for i in folder_list:
        
        if new_path == ' ':
            path = f'{i}/'
        else: 
            path = new_path + f'{i}/'

        if not os.path.exists(path):
            os.mkdir(path)
            
        new_path = path
    return new_path

def remove_file(file_path: str):
    """
        file_path: Enter the file path you want to delete

        returns the string value of new path created
    """

    if os.path.exists(file_path): os.remove(file_path)


def move_file(source_path: str, destination_path: str):
    """
        source_path : path/file
        destination_path : path/file
    """
    source = f'{source_path}'
    destination = f'{destination_path}'
    remove_file(destination)
    shutil.move(source, destination)
    
def create_folder(test_name: str):
    if not os.path.exists('waveforms'): os.mkdir('waveforms')
    waveforms_folder = f'waveforms/{test_name}'
    pathname = f"{os.getcwd()}/{waveforms_folder}"
    if not os.path.exists(pathname): os.mkdir(pathname)
    

def clear_sheet(directory: str,filename: str,sheet_name: str):
    """Clears the specified sheet from the workbook if it exists"""
    dst = f"{directory}/{filename}.xlsx"
    if os.path.exists(dst):
        wb: Workbook = load_workbook(dst)
        sheet_list = wb.sheetnames  
        if sheet_name in sheet_list:
            wb.remove(wb[sheet_name])
            wb.create_sheet(sheet_name)
            wb.save(dst)
        wb.close()
        
def get_anchor(col:int, row:int):
    """
    get anchor given a numerical col row  -> (col = 2, row = 4 -> 'B4')
    returns anchor (str)
    """
    anchor = f"{get_column_letter(col)}{row}"
    return anchor

def col_row_extractor(excel_coordinate: str):
    """
    extract col and row given an excel coordinate (i.e. 'B4' -> col = 2, row = 4)

    excel_coordinate : i.e. 'B4' (str)
    returns col, row (int)
    """
    coordinates = coordinate_from_string(excel_coordinate)
    col = column_index_from_string(coordinates[0])
    row = coordinates[1]
    return col, row

def excel_to_df(filename: str, sheet_name: str, start_corner: str, end_corner: str):
    """
    reading dataframe from excel.
    
    filename     : must include full filename path (cwd + path + file.extension)
    sheet_name   : sheet name in excel file
    start_corner : cell coordinate to start selection of data
    end_corner   : cell coordinate to end selection of data

    returns df
    """

    # print(f"reading dataframe from {filename} {sheet_name}")

    start_col, start_row = col_row_extractor(start_corner)
    end_col, end_row = col_row_extractor(end_corner)

    skiprows = start_row - 2
    usecols = f'{get_column_letter(start_col)}:{get_column_letter(end_col)}'
    nrows = end_row - start_row + 1

    return pd.read_excel(filename, sheet_name, skiprows=skiprows, usecols=usecols, nrows=nrows)
    # return pd.read_csv(filename, sheet_name, skiprows=skiprows, usecols=usecols, nrows=nrows)

def df_to_excel(wb: Workbook, sheet_name: str, df: pd.DataFrame, anchor: str):
    """
    writing dataframe to excel.

    wb          : workbook
    sheet_name  : sheet name in excel file
    df          : dataframe
    anchor      : anchor point in excel

    returns None
    """



    sheet_list = wb.sheetnames
    if sheet_name not in sheet_list: 
        wb.create_sheet(sheet_name)
    #else: 
        #wb.remove_sheet(wb.get_sheet_by_name(sheet_name))
        #wb.create_sheet(sheet_name)
    try:
        default_sheet = wb['Sheet']
        wb.remove(default_sheet)
    except: pass

    start_col, start_row = col_row_extractor(anchor)
    df_row_len, df_col_len = df.shape
    end_row = start_row + df_row_len - 1
    end_col = start_col + df_col_len - 1

    for row in range(start_row, end_row+1):
        for col in range(start_col, end_col+1):
            wb[sheet_name][f'{get_column_letter(col)}{row}'] = df.iloc[row-start_row, col-start_col]
            wb[sheet_name][f'{get_column_letter(col)}{row}'].alignment = Alignment(horizontal='center',vertical='center')
            
            if ((col%14) in [9, 10]) & (row > 5):
                wb[sheet_name][f'{get_column_letter(col)}{row}'].number_format = '#,##0.000'
            elif ((((col%14) in  [4, 5, 6, 7, 8 ,11 ,12, 13]) |  (((col%14) +14 in [14, 15, 16, 17]) & col > 3)) & (row > 5)):
                wb[sheet_name][f'{get_column_letter(col)}{row}'].number_format= '#,##0.00'
                
def image_to_excel(wb: Workbook, sheet_name: str, filename: str, folder_path: str, anchor: str):
    """
    writing image to excel.

    image size -> 39 rows, 16 columns
    wb          : workbook
    sheet_name  : sheet name in excel file
    filename    : filename of theh image
    folder_path : image location
    anchor      : anchor point in excel
    """

    file = folder_path + filename
    # file = os.getcwd() + folder_path + filename
    img = openpyxl.drawing.image.Image(file)
    img.anchor = anchor
    img.width = 1056
    img.height = 659.90551181

    sheet_list = wb.sheetnames
    if sheet_name not in sheet_list: wb.create_sheet(sheet_name)
    #else: 
        #wb.remove_sheet(wb.get_sheet_by_name(sheet_name))
        #wb.create_sheet(sheet_name)
    try:
        default_sheet = wb['Sheet']
        wb.remove(default_sheet)
    except: pass

    col, row = col_row_extractor(anchor)
    wb[sheet_name][f'{get_column_letter(col)}{row-1}'] = filename
    wb[sheet_name].add_image(img)

def change_number_format(wb: Workbook, sheet_name: str ,start_row: int, start_col:int , end_row:int, end_col:int,num_format: str):
    """
    change number formatting on a specific range

    wb          : workbook, must be openpyxl.workbook.Workbook object
    sheet_name  : sheet name in excel file, string input
    start_row   : starting row of range, integer input
    start_col   : starting column of range, integer input
    start_col   : ending row of range, integer input
    start_col   : ending column of  range, integer input
    num_format  : format of number, string input (eg. "###0.000")
    """
    
    ws = wb[sheet_name]
    for row in range(start_row, end_row+1):
        for col in range(start_col, end_col+1):
            ws[f'{get_column_letter(col)}{row}'].number_format = num_format

def copy_sheet(wb: Workbook, wbs: Workbook, sheet_name: str):
    """
	copy worksheet from one workbook to another
	wb			: target workbook, must be openpyxl.workbook.Workbook object
	wbs			: source workbook, must be openpyxl.workbook.Workbook object
	sheet_name	: name of sheet to be copied, string input
	"""
    wb.create_sheet(sheet_name)
    
    for (row, col), source_cell  in wbs[sheet_name]._cells.items():
        target_cell =  wb[sheet_name].cell(column=col, row=row)

        target_cell._value = source_cell._value
        target_cell.data_type = source_cell.data_type
        
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)

        if source_cell.hyperlink:
            target_cell._hyperlink = copy(source_cell.hyperlink)

        if source_cell.comment:
            target_cell.comment = copy(source_cell.comment)
            
        target_cell.alignment = Alignment(wrap_text=False, vertical='center',horizontal='center') 
        
    wb[sheet_name].sheet_format = copy(wbs[sheet_name].sheet_format)
    wb[sheet_name].sheet_properties = copy(wbs[sheet_name].sheet_properties)
    wb[sheet_name].merged_cells = copy(wbs[sheet_name].merged_cells)
    wb[sheet_name].page_margins = copy(wbs[sheet_name].page_margins)
    wb[sheet_name].page_setup = copy(wbs[sheet_name].page_setup)
    wb[sheet_name].print_options = copy(wbs[sheet_name].print_options)
    
    return wb
            
#################### DATA EXPORT CODE ###########################################################
def dataframe_from_headers(df_header_list):
    df = pd.DataFrame(columns = df_header_list)
    df.loc[len(df)] = df_header_list
    return df
# @timeit
def export_to_excel(df: pd.DataFrame, output_folder_path: str, excel_name: str, sheet_name: str, anchor: str):
    #df.loc[len(df)] = output_list
   #print(output_list)


    dst = f"{output_folder_path}/{excel_name}.xlsx"
    if not os.path.exists(dst): 
        #shutil.copyfile(src, dst)
        wb = openpyxl.Workbook()
        wb.save(dst)
    wb = load_workbook(dst)
    df_to_excel(wb, sheet_name, df, anchor)
    wb.save(dst)
    wb.close()

def export_screenshot_to_excel(excel_name: str, waveforms_folder: str, sheet_name: str, filename: str, anchor: str):

    src = f"{os.getcwd()}/blank.xlsx"
    dst = f"{waveforms_folder}/{excel_name}.xlsx"
    if not os.path.exists(dst): shutil.copyfile(src, dst)

    wb = load_workbook(dst)
    image_to_excel(wb, sheet_name, filename=filename, folder_path=waveforms_folder, anchor=anchor)
    wb.save(dst)
    wb.close()


def generate_plots_LoadReg(
    vout: float,
    iout: float,
    num_step: int,
    vin_step: int,
    coupling: str,
    wb: Workbook,
    sheet_name: str,
    wb_filepath: str,
    usb_pd_flag: bool):
    
    """
    vout            : output voltage setpoint, numerical input in V
    iout            : nominal output current, numerical input in A
    num_step        : number of load step per sheet, integer input
    vin_step        : number of line input step per sheet, integer input
    coupling        : input coupling, 'AC' or 'DC'
    wb              : workbook containing the data, must be openpyxl.workbook.Workbook object
    sheet_name      : sheet within the workbook containing the data, string input
    wb_filepath     : path and filename of excel file where charts will be saved, string input
    usb_pd_flag     : for 5V USB-PD condition, limits are 4.75 V and 5.5 V instead of 5%, boolean input
    """
    
    template_sheet_name = 'Charts_Template'
    
    # Check the files , will raise Exception if not
    check_if_files_exist(
        template_sheet_name = template_sheet_name,
        datasrc_sheet_name = sheet_name,
        datasrc_wb_filepath = wb_filepath)
    
    # Load both the template workbook and data source workbook
    wbs = load_workbook(chart_template_path)
    wb = load_workbook(wb_filepath)

    # Get the sheet containing the chart template
    ws_template = wbs[template_sheet_name]
    
    # Reset chartsheet and get the worksheet object
    chart_sheet_name = f"LoadReg_{coupling}_{vout:g}V_{iout:g}A"
    ws = reset_chartsheet(wb, chart_sheet_name)

    # Copy chart from template then close template workbook
    chart = deepcopy(ws_template._charts[0])
    wbs.close()     
    
    #Initialize chart layout properties
    
    xLabel = "Load (A)"
    yLabel = "Output Voltage (V)"
    
    if vout <= 5:
        y_min_scale = math.floor(vout-2)
        y_max_scale = math.ceil(vout+2)
        y_major_unit = 0.5
        y_numFmt = '0.0'
    else:
        y_min_scale = math.floor(vout-4)
        y_max_scale = math.ceil(vout+4)
        y_major_unit = 1
        y_numFmt = '0'
    
    x_min_scale = 0
    x_max_scale = round(iout,6)
    
    if iout <= 0.5:
        x_major_unit = 0.05
        x_numFmt = '0.00'
    elif iout <= 1:
        x_major_unit = 0.1
        x_numFmt = '0.0'
    elif iout <= 3:
        x_major_unit = 0.3
        x_numFmt = '0.0'
    else:
        x_major_unit = 0.5
        x_numFmt = '0.0'
        

    chart = chart_setup(chart," ",xLabel,yLabel,
                x_min_scale,x_max_scale,y_min_scale,y_max_scale,x_major_unit,
                0.1,y_major_unit,0.1,x_numFmt,y_numFmt)
    
    #Add chart series per line voltage input
        
    for series_index in range(vin_step):
        series_title = f"{wb[sheet_name][f'B{6+(num_step+4)*series_index}'].value} V{coupling.upper()}"
        xvalues = Reference(wb[sheet_name], min_col=10, min_row=6+((num_step+4)*series_index), max_row=(6+(num_step-1)+((num_step+4)*series_index)))
        values = Reference(wb[sheet_name], min_col=9, min_row=6+((num_step+4)*series_index), max_row=(6+(num_step-1)+((num_step+4)*series_index)))
        if series_index <= 3:
            chart.series[series_index] = edit_series(chart.series[series_index],xvalues,values,series_title)
        else:
            series = deepcopy(chart.series[series_index-1])
            chart.series.append(series)
            chart.series[series_index].graphicalProperties.line.solidFill = color_list[series_index-4]
            chart.series[series_index] = edit_series(chart.series[series_index],xvalues,values,series_title)
            chart.series[series_index].marker.spPr.solidFill = color_list[series_index-4]
            chart.series[series_index].marker.spPr.ln = LineProperties(noFill=True)
    series_index = series_index + 1
    temp = series_index

    #Add buffer value to proceeding series if vin_step < 4 in order to deepcopy last series and be able to edit its value

    while len(chart.series) > (temp):
        chart.series[temp] = edit_series(chart.series[temp],xvalues,values,series_title)
        temp = temp + 1
        
    #Set load regulation limits to +-5%
    #For USB PD devices, 5V limits set to 4.75 V & 5.5 V    
    
    if (vout == 5) & (usb_pd_flag == True):
        upper_limit = 5.5
        lower_limit = 4.75
        ul_title ='5.5 V '
        ll_title = '4.75 V'
    else:
        upper_limit = vout*1.05
        lower_limit = vout*0.95
        ul_title = '+5%'
        ll_title = '-5%'

    
    #Add chart series for limits
    
    series = deepcopy(chart.series[len(chart.series) -1])
    chart.series.append(series)
    xvalues = f"={{{x_min_scale},{x_max_scale}}}"
    values = f"={{{upper_limit},{upper_limit}}}"
    chart.series[len(chart.series)-1].xVal.numRef.f = xvalues
    chart.series[len(chart.series)-1].yVal.numRef.f = values
    chart.series[len(chart.series)-1].title.v = ul_title
    chart.series[len(chart.series)-1].graphicalProperties.line.solidFill = "92D050"
    chart.series[len(chart.series)-1].graphicalProperties.line.dashStyle = "dash"
    chart.series[len(chart.series)-1].marker = openpyxl.chart.marker.Marker(None)
    
    series = deepcopy(chart.series[len(chart.series) -1])
    chart.series.append(series)
    xvalues = f"{{{x_min_scale},{x_max_scale}}}"
    values = f"{{{lower_limit},{lower_limit}}}"
    chart.series[len(chart.series)-1].xVal.numRef.f = xvalues
    chart.series[len(chart.series)-1].yVal.numRef.f = values
    chart.series[len(chart.series)-1].title.v = ll_title
    chart.series[len(chart.series)-1].graphicalProperties.line.solidFill = "92D050"
    chart.series[len(chart.series)-1].graphicalProperties.line.dashStyle = "dash"
    chart.series[len(chart.series)-1].marker = openpyxl.chart.marker.Marker(None)
    
    #Remove extra chart series (if vin_step < 4)
    
    while len(chart.series) > (series_index+2):
        chart.series.pop(len(chart.series)-3)
        
    
    #Clear chart title then place to chartsheet and save workbook
    
    chart.title = " "  
    ws.add_chart(chart)        
    wb.save(wb_filepath)
    return wb

def generate_plots_LineReg(
    vout: float,
    iout: float,
    num_step: int,
    vin_step: int,
    coupling: str,
    wb: Workbook,
    sheet_name: str,
    wb_filepath: str,
    usb_pd_flag: bool):
    
    """
    vout            : output voltage setpoint, numerical input in V
    iout            : nominal output current, numerical input in A
    num_step        : number of load step per sheet, integer input
    vin_step        : number of line input step per sheet, integer input
    coupling        : coupling of input voltage, either "DC" or "AC"
    wb              : workbook containing the data, must be openpyxl.workbook.Workbook object
    sheet_name      : sheet within the workbook containing the data, string input
    wb_filepath     : path and filename of excel file where charts will be saved, string input
    usb_pd_flag     : for 5V USB-PD condition, limits are 4.75 V and 5.5 V instead of 5%, boolean input
    """
    
    template_sheet_name = 'Charts_Template_LineReg'
    
    # Check the files , will raise Exception if not
    check_if_files_exist(
        template_sheet_name = template_sheet_name,
        datasrc_sheet_name = sheet_name,
        datasrc_wb_filepath = wb_filepath)
    
    # Load both the template workbook and data source workbook
    wbs = load_workbook(chart_template_path)
    wb = load_workbook(wb_filepath)

    # Get the sheet containing the chart template
    ws_template = wbs[template_sheet_name]
    
    # Reset chartsheet and get the worksheet object
    chart_sheet_name = f"LineReg_{coupling}_{vout:g}V_{iout:g}A"
    ws = reset_chartsheet(wb, chart_sheet_name)

    # Copy chart from template then close template workbook
    chart = deepcopy(ws_template._charts[0])
    wbs.close()    
    
    #Initialize chart layout properties
    
    xLabel = f"Input Voltage (V{coupling.upper()})"
    yLabel = "Output Voltage (V)"
    
    if vout <= 5:
        y_min_scale = math.floor(vout-2)
        y_max_scale = math.ceil(vout+2)
        y_major_unit = 0.5
        y_numFmt = '0.0'
    else:
        y_min_scale = math.floor(vout-4)
        y_max_scale = math.ceil(vout+4)
        y_major_unit = 1
        y_numFmt = '0'
        
    x_min_scale = (math.floor((wb[sheet_name]['B6'].value)/10))*10-5
    x_max_scale = (math.ceil((wb[sheet_name][f'B{6+vin_step-1}'].value)/20))*20+5
    x_major_unit = 20
    x_numFmt = '0'

    chart = chart_setup(chart," ",xLabel,yLabel,
                x_min_scale,x_max_scale,y_min_scale,y_max_scale,x_major_unit,
                0.1,y_major_unit,0.1,x_numFmt,y_numFmt)
    
    #Add chart series per line voltage input
        
    for series_index in range(num_step):
        series_title = f"{wb[sheet_name][f'A{6+(vin_step+4)*series_index}'].value}% Load"
        xvalues = Reference(wb[sheet_name], min_col=4, min_row=6+((vin_step+4)*series_index), max_row=(6+(vin_step-1)+((vin_step+4)*series_index)))
        values = Reference(wb[sheet_name], min_col=9, min_row=6+((vin_step+4)*series_index), max_row=(6+(vin_step-1)+((vin_step+4)*series_index)))
        if series_index <= 3:
            chart.series[series_index] = edit_series(chart.series[series_index],xvalues,values,series_title)
        else:
            series = deepcopy(chart.series[series_index-1])
            chart.series.append(series)
            chart.series[series_index].graphicalProperties.line.solidFill = color_list[series_index-4]
            chart.series[series_index] = edit_series(chart.series[series_index],xvalues,values,series_title)
            chart.series[series_index].marker.spPr.solidFill = color_list[series_index-4]
            chart.series[series_index].marker.spPr.ln = LineProperties(noFill=True)
    series_index = series_index + 1
    temp = series_index

    #Add buffer value to proceeding series if vin_step < 4 in order to deepcopy last series and be able to edit its value

    while len(chart.series) > (temp):
        chart.series[temp] = edit_series(chart.series[temp],xvalues,values,series_title)
        temp = temp + 1
        
    #Set load regulation limits to +-5%
    #For USB PD devices, 5V limits set to 4.75 V & 5.5 V    
    
    if (vout == 5) & (usb_pd_flag == True):
        upper_limit = 5.5
        lower_limit = 4.75
        ul_title ='5.5 V '
        ll_title = '4.75 V'
    else:
        upper_limit = vout*1.05
        lower_limit = vout*0.95
        ul_title = '+5%'
        ll_title = '-5%'

    
    #Add chart series for limits
    
    series = deepcopy(chart.series[len(chart.series) -1])
    chart.series.append(series)
    xvalues = f"={{{x_min_scale},{x_max_scale}}}"
    values = f"={{{upper_limit},{upper_limit}}}"
    chart.series[len(chart.series)-1].xVal.numRef.f = xvalues
    chart.series[len(chart.series)-1].yVal.numRef.f = values
    chart.series[len(chart.series)-1].title.v = ul_title
    chart.series[len(chart.series)-1].graphicalProperties.line.solidFill = "92D050"
    chart.series[len(chart.series)-1].graphicalProperties.line.dashStyle = "dash"
    chart.series[len(chart.series)-1].marker = openpyxl.chart.marker.Marker(None)
    
    series = deepcopy(chart.series[len(chart.series) -1])
    chart.series.append(series)
    xvalues = f"{{{x_min_scale},{x_max_scale}}}"
    values = f"{{{lower_limit},{lower_limit}}}"
    chart.series[len(chart.series)-1].xVal.numRef.f = xvalues
    chart.series[len(chart.series)-1].yVal.numRef.f = values
    chart.series[len(chart.series)-1].title.v = ll_title
    chart.series[len(chart.series)-1].graphicalProperties.line.solidFill = "92D050"
    chart.series[len(chart.series)-1].graphicalProperties.line.dashStyle = "dash"
    chart.series[len(chart.series)-1].marker = openpyxl.chart.marker.Marker(None)
    
    #Remove extra chart series (if vin_step < 4)
    
    while len(chart.series) > (series_index+2):
        chart.series.pop(len(chart.series)-3)
        
    
    #Clear chart title then place to chartsheet and save workbook
    
    chart.title = " "  
    ws.add_chart(chart)        
    wb.save(wb_filepath)
    return wb

def generate_plots_LoadvEff(
    vout: float,
    iout: float,
    num_step: int,
    vin_step: int,
    coupling: str,
    wb: Workbook,
    sheet_name: str,
    wb_filepath: str):
    
    """
    vout            : output voltage setpoint, numerical input in V
    iout            : nominal output current, numerical input in A
    num_step        : number of load step per sheet, integer input
    vin_step        : number of line input step per sheet, integer input
    coupling        : input coupling, 'AC' or 'DC'
    wb              : workbook containing the data, must be openpyxl.workbook.Workbook object
    sheet_name      : sheet within the workbook containing the data, string input
    wb_filepath     : path and filename of excel file where charts will be saved, string input
    """
       
    template_sheet_name = 'Charts_Template'
    
    # Check the files , will raise Exception if not
    check_if_files_exist(
        template_sheet_name = template_sheet_name,
        datasrc_sheet_name = sheet_name,
        datasrc_wb_filepath = wb_filepath)
    
    # Load both the template workbook and data source workbook
    wbs = load_workbook(chart_template_path)
    wb = load_workbook(wb_filepath)

    # Get the sheet containing the chart template
    ws_template = wbs[template_sheet_name]
    
    # Reset chartsheet and get the worksheet object
    chart_sheet_name = f"LdvEff_{coupling}_{vout:g}V_{iout:g}A" 
    ws = reset_chartsheet(wb, chart_sheet_name)

    # Copy chart from template then close template workbook
    chart = deepcopy(ws_template._charts[0])
    wbs.close()     
    
    #Initialize chart layout properties
    
    xLabel = "Load (A)"
    yLabel = "Efficiency (%)"
                                                                                      
    y_min_scale = 50
    y_max_scale = 100
    y_major_unit = 5
    y_numFmt = '0'
    x_min_scale = 0
    x_max_scale = round(iout,6)
    if iout <= 0.5:
        x_major_unit = 0.05
        x_numFmt = '0.00'
    elif iout <= 1:
        x_major_unit = 0.1
        x_numFmt = '0.0'
    elif iout <= 3:
        x_major_unit = 0.3
        x_numFmt = '0.0'
    else:
        x_major_unit = 0.5
        x_numFmt = '0.0'
    
    chart = chart_setup(chart," ",xLabel,yLabel,
                x_min_scale,x_max_scale,y_min_scale,y_max_scale,x_major_unit,
                1,y_major_unit,1,x_numFmt,y_numFmt)
    
    #Add chart series per line voltage input
    for series_index in range(vin_step):
        
        #to check if 0A in either start or end of the series
        iout_ref = wb[sheet_name][f'J{6+(num_step+4)*series_index}'].value
        if abs(iout_ref) < 0.005:
            series_offset = 1
        else:
            series_offset = 0
            
        iout_ref = wb[sheet_name][f'J{6+(num_step-1)+((num_step+4)*series_index)}'].value
        if abs(iout_ref) < 0.005:
            series_trim = 1
        else:
            series_trim = 0
            
        series_title = f"{wb[sheet_name][f'B{6+(num_step+4)*series_index}'].value} V{coupling.upper()}"
        xvalues = Reference(wb[sheet_name], min_col=10, min_row=6+series_offset+((num_step+4)*series_index), max_row=6+(num_step-1-series_trim)+((num_step+4)*series_index))
        values = Reference(wb[sheet_name], min_col=13, min_row=6+series_offset+((num_step+4)*series_index), max_row=6+(num_step-1-series_trim)+((num_step+4)*series_index))
        if series_index <= 3:
            chart.series[series_index] = edit_series(chart.series[series_index],xvalues,values,series_title)
        else:
            series = deepcopy(chart.series[series_index-1])
            chart.series.append(series)
            chart.series[series_index].graphicalProperties.line.solidFill = color_list[series_index-4]
            chart.series[series_index] = edit_series(chart.series[series_index],xvalues,values,series_title)
            chart.series[series_index].marker.spPr.solidFill = color_list[series_index-4]
            chart.series[series_index].marker.spPr.ln = LineProperties(noFill=True)
    series_index = series_index + 1
    while len(chart.series) > series_index:
        chart.series.pop()
       
    #Clear chart title then place to chartsheet and save workbook
    
    chart.title = " "
    ws.add_chart(chart)       
    wb.save(wb_filepath)
    return wb
    
def generate_plots_LinevEff(
    vout: float,
    iout: float,
    num_step: int,
    vin_step: int,
    coupling: str,
    wb: Workbook,
    sheet_name: str,
    wb_filepath: str):
    
    """
    vout            : output voltage setpoint, numerical input
    iout            : output current setpoint, numerical input
    num_step        : number of load step per sheet, integer input
    vin_step        : number of line input step per sheet, integer input
    coupling        : coupling of input voltage, either "DC" or "AC"
    wb              : workbook containing the data, must be openpyxl.workbook.Workbook object
    sheet_name      : sheet within the workbook containing the data, string input
    wb_filepath     : path and filename of excel file where charts will be saved, string input
    """
    
    template_sheet_name = 'Charts_Template_LineReg'
    
    # Check the files , will raise Exception if not
    check_if_files_exist(
        template_sheet_name = template_sheet_name,
        datasrc_sheet_name = sheet_name,
        datasrc_wb_filepath = wb_filepath)
    
    # Load both the template workbook and data source workbook
    wbs = load_workbook(chart_template_path)
    wb = load_workbook(wb_filepath)

    # Get the sheet containing the chart template
    ws_template = wbs[template_sheet_name]
    
    # Reset chartsheet and get the worksheet object
    chart_sheet_name = f"LnvEff_{coupling}_{vout:g}V_{iout:g}A" 
    ws = reset_chartsheet(wb, chart_sheet_name)

    # Copy chart from template then close template workbook
    chart = deepcopy(ws_template._charts[0])
    wbs.close()     
    
    #Initialize chart layout properties
    
    xLabel = f"Input Voltage (V{coupling.upper()})"
    yLabel = "Efficiency (%)"
    
    y_min_scale = 50
    y_max_scale = 100
    y_major_unit = 5
    y_numFmt = '0'
    x_min_scale = (math.floor((wb[sheet_name]['B6'].value)/10))*10-5
    x_max_scale = (math.ceil((wb[sheet_name][f'B{6+vin_step-1}'].value)/20))*20+5
    x_major_unit = 20
    x_numFmt = '0'
    
    chart = chart_setup(chart," ",xLabel,yLabel,
                x_min_scale,x_max_scale,y_min_scale,y_max_scale,x_major_unit,
                1,y_major_unit,1,x_numFmt,y_numFmt)
    
    #Add chart series per line voltage input
        
    for series_index in range(num_step):
        series_title = f"{wb[sheet_name][f'A{6+(vin_step+4)*series_index}'].value}% Load"
        xvalues = Reference(wb[sheet_name], min_col=4, min_row=6+((vin_step+4)*series_index), max_row=(6+(vin_step-1)+((vin_step+4)*series_index)))
        values = Reference(wb[sheet_name], min_col=13, min_row=6+((vin_step+4)*series_index), max_row=(6+(vin_step-1)+((vin_step+4)*series_index)))
        if series_index <= 3:
            chart.series[series_index] = edit_series(chart.series[series_index],xvalues,values,series_title)
        else:
            series = deepcopy(chart.series[series_index-1])
            chart.series.append(series)
            chart.series[series_index].graphicalProperties.line.solidFill = color_list[series_index-4]
            chart.series[series_index] = edit_series(chart.series[series_index],xvalues,values,series_title)
            chart.series[series_index].marker.spPr.solidFill = color_list[series_index-4]
            chart.series[series_index].marker.spPr.ln = LineProperties(noFill=True)
    series_index = series_index + 1
    
    #Remove extra chart series (if vin_step < 4)
    
    while len(chart.series) > (series_index):
        chart.series.pop()
    
    #Clear chart title then place to chartsheet and save workbook
    
    chart.title = " "  
    ws.add_chart(chart)        
    wb.save(wb_filepath)
    return wb

def generate_plots_LinevEff2(
    vout: float,
    num_step: int,
    vin_step: int,
    coupling: str,
    wb: Workbook,
    sheet_name_list: list,
    wb_filepath: str):
    
    """
    vout            : list of output voltage setpoint, numerical list input
    num_step        : number of load step per sheet, integer input
    vin_step        : number of line input step per sheet, integer input
    coupling        : coupling of input voltage, either "DC" or "AC"
    wb              : workbook containing the data, must be openpyxl.workbook.Workbook object
    sheet_list      : list of sheet within the workbook containing the data corresponding to the items in vout, list of string input
    wb_filepath     : path and filename of excel file where charts will be saved, string input
    """
    # Check if files and folders exist
    template_sheet_name = 'Charts_Template_Eff_v_Line'
    
    # Check the files , will raise Exception if not
    check_if_files_exist(
        template_sheet_name = template_sheet_name,
        datasrc_sheet_name = sheet_name_list[0],
        datasrc_wb_filepath = wb_filepath)
    
    # Load both the template workbook and data source workbook
    wbs = load_workbook(chart_template_path)
    wb = load_workbook(wb_filepath)

    # Get the sheet containing the chart template
    ws_template = wbs[template_sheet_name]
    
    
    # Loop for each loadstep
    for load_index in range(num_step):
    
        Load = wb[sheet_name_list[0]][f'A{6+(vin_step+4)*load_index}'].value
        # Reset chartsheet and get the worksheet object
        chart_sheet_name = f'Charts_LinevEff_{round(Load,3):g}%'
        ws = reset_chartsheet(wb, chart_sheet_name)
        
        # Copy chart from template then close template workbook
        chart = deepcopy(ws_template._charts[0])

        # Initialize chart layout properties

        xLabel = f"Input Voltage (V{coupling.upper()})"
        yLabel = "Efficiency (%)"

        y_min_scale = 50
        y_max_scale = 100
        y_major_unit = 5
        y_numFmt = '0'
        x_min_scale = (math.floor((wb[sheet_name_list[0]]['B6'].value)/10))*10-5
        x_max_scale = (math.ceil((wb[sheet_name_list[0]][f'B{6+vin_step-1}'].value)/20))*20+5
        x_major_unit = 20
        x_numFmt = '0'

        chart = chart_setup(chart," ",xLabel,yLabel,
                    x_min_scale,x_max_scale,y_min_scale,y_max_scale,x_major_unit,
                    1,y_major_unit,1,x_numFmt,y_numFmt)
        # Loop for each output voltage
        for sheet_index, sheet in enumerate(sheet_name_list):
            if sheet not in wb:
                print(f'sheet "{sheet}" not in workbook')
                return
            Iout = round(wb[sheet][f'J{6+(vin_step+4)*load_index}'].value,2)
            if round(100*(round(Iout,2)-round(Iout,1))) != 0:      # for rounding off value to 2 decimals or up to the nearest nonzero decimal 
                Iout = round(Iout,2)
            elif round(10*(round(Iout,1)-round(Iout))) != 0: 
                Iout = round(Iout,1)
            else:
                Iout = round(Iout)
            # Add chart series per voltage output

            series_title = f'{vout[sheet_index]}V / {Iout}A'
            xvalues = Reference(wb[sheet], min_col=4, min_row=6+(vin_step+4)*load_index, max_row=6+(vin_step+4)*load_index+(vin_step-1))
            values = Reference(wb[sheet], min_col=13, min_row=6+(vin_step+4)*load_index, max_row=6+(vin_step+4)*load_index+(vin_step-1))
            if sheet_index <= 3:
                chart.series[sheet_index] = edit_series(chart.series[sheet_index],xvalues,values,series_title)
            else:
                series = deepcopy(chart.series[sheet_index-1])
                chart.series.append(series)
                chart.series[sheet_index].graphicalProperties.line.solidFill = color_list[sheet_index-4]
                chart.series[sheet_index] = edit_series(chart.series[sheet_index],xvalues,values,series_title)
                chart.series[sheet_index].marker.spPr.solidFill = color_list[sheet_index-4]
                chart.series[sheet_index].marker.spPr.ln = LineProperties(noFill=True)

        # Remove extra chart series (if vin_step < 4)

        while len(chart.series) > (len(sheet_name_list)):
            chart.series.pop()

        # Clear chart title then place to chartsheet and save workbook

        chart.title = " "  
        ws.add_chart(chart)       
        
    wb.save(wb_filepath)    
    wbs.close()
    return wb

def generate_plots_LoadvRipple(
    vout: float,
    iout: float,
    num_step: int,
    vin_step: int,
    coupling: str,
    wb: Workbook,
    sheet_name: str,
    wb_filepath: str):
    
    """
    vout            : output voltage setpoint, numerical input in V
    iout            : nominal output current, numerical input in A
    num_step        : number of load step per sheet, integer input
    vin_step        : number of line input step per sheet, integer input
    coupling        : input coupling, 'AC' or 'DC'
    wb              : workbook containing the data, must be openpyxl.workbook.Workbook object
    sheet_name      : sheet within the workbook containing the data, string input
    wb_filepath     : path and filename of excel file where charts will be saved, string input
    """
    # Check if files and folders exist
    template_sheet_name = 'Charts_Template'
    
    # Check the files , will raise Exception if not
    check_if_files_exist(
        template_sheet_name = template_sheet_name,
        datasrc_sheet_name = sheet_name,
        datasrc_wb_filepath = wb_filepath)
    
    # Load both the template workbook and data source workbook
    wbs = load_workbook(chart_template_path)
    wb = load_workbook(wb_filepath)

    # Get the sheet containing the chart template
    ws_template = wbs[template_sheet_name]
    
    # Reset chartsheet and get the worksheet object
    chart_sheet_name = f"LdvRipple_{coupling}_{vout:g}V_{iout:g}A" 
    ws = reset_chartsheet(wb, chart_sheet_name)

    # Copy chart from template then close template workbook
    chart = deepcopy(ws_template._charts[0])
    wbs.close() 
    
    #Initialize chart layout properties
    
    xLabel = "Load (A)"
    yLabel = "Output Ripple (mV)"
    data_ptp = excel_to_df(wb_filepath,sheet_name,'O6',f'O{6+(num_step-1)+((num_step+4)*(vin_step-1))}')
    y_min_scale = 0
    y_max_scale = (math.ceil(max(data_ptp.values)/50)+1)*50
    if y_max_scale < 200:
        y_max_scale = 200
    y_major_unit = y_max_scale/10
    y_numFmt = '0'
    x_min_scale = 0
    x_max_scale = round(iout,6)
    if iout <= 0.5:
        x_major_unit = 0.05
        x_numFmt = '0.00'
    elif iout <= 1:
        x_major_unit = 0.1
        x_numFmt = '0.0'
    elif iout <= 3:
        x_major_unit = 0.3
        x_numFmt = '0.0'
    else:
        x_major_unit = 0.5
        x_numFmt = '0.0'

    chart = chart_setup(chart," ",xLabel,yLabel,
                x_min_scale,x_max_scale,y_min_scale,y_max_scale,x_major_unit,
                1,y_major_unit,1,x_numFmt,y_numFmt)
    
    #Add chart series per line voltage input
    
    for series_index in range(vin_step):
        series_title = f"{wb[sheet_name][f'B{6+(num_step+4)*series_index}'].value} V{coupling.upper()}"
        xvalues = Reference(wb[sheet_name], min_col=10, min_row=6+((num_step+4)*series_index), max_row=6+(num_step-1)+((num_step+4)*series_index))
        values = Reference(wb[sheet_name], min_col=15, min_row=6+((num_step+4)*series_index), max_row=6+(num_step-1)+((num_step+4)*series_index))
        if series_index <= 3:
            chart.series[series_index] = edit_series(chart.series[series_index],xvalues,values,series_title)
        else:
            series = deepcopy(chart.series[series_index-1])
            chart.series.append(series)
            chart.series[series_index].graphicalProperties.line.solidFill = color_list[series_index-4]
            chart.series[series_index] = edit_series(chart.series[series_index],xvalues,values,series_title)
            chart.series[series_index].marker.spPr.solidFill = color_list[series_index-4]
            chart.series[series_index].marker.spPr.ln = LineProperties(noFill=True)
    series_index = series_index + 1
    while len(chart.series) > series_index:
        chart.series.pop()
    
    #Clear chart title then place to chartsheet and save workbook
    
    chart.title = " "    
    
    ws.add_chart(chart)
    wb.save(wb_filepath)
    return wb


def check_if_files_exist(
        template_sheet_name, 
        datasrc_sheet_name, 
        datasrc_wb_filepath):
    # Check if the template file exists
    if not os.path.exists(chart_template_path):
        raise ExcelPlotterError(f'Charts template file does not exist')
    
    # Check if the data source workbook exists
    if not os.path.exists(datasrc_wb_filepath):
        raise ExcelPlotterError(f'Data source workbook file does not exist')

    # Load the template workbook
    wbs = load_workbook(chart_template_path)
    if template_sheet_name not in wbs.sheetnames:
        raise ExcelPlotterError(f'{template_sheet_name} does not exist')
    
    # Check if the charts template exists
    ws_template = wbs[template_sheet_name]
    if len(ws_template._charts) == 0:
        raise ExcelPlotterError('Charts template does not exist')
    
    # Check if the specified sheet exists
    wb = load_workbook(datasrc_wb_filepath)
    if datasrc_sheet_name not in wb:
        raise ExcelPlotterError(f'sheet "{datasrc_sheet_name}" not in workbook')
    
def reset_chartsheet(wb:Workbook, sheet_name):
    """Resets the specified chartsheet if it exists.
    Create a new chart sheet if it doesn't.
    Return the worksheet object"""
    if f"Charts_{sheet_name}" in wb.sheetnames:
        wb.remove(wb[f"Charts_{sheet_name}"])
    wb.create_chartsheet(f"Charts_{sheet_name}",0)
    ws = wb[f"Charts_{sheet_name}"]
    
    return ws

class CVCCPlot():

    template_sheet_name = 'Charts_Template'
    # col_vout = 7
    # col_iout = 8
    row_startdata=6

    def __init__(self,
                 vout_v: float,
                 iout_a: float,
                 vin_step: int,
                 coupling: str,
                 num_step: int,
                 sheet_name: str,
                 wb_filepath: str,
                 column_step: int,
                 col_vout: int,
                 col_iout: int,
                 cc_current_start_A: float = None,
                 cc_voltage_start_V: float = None,
                 cc_current_end_A: float = None,
                 cc_voltage_end_V: float = None,
                 x_axis_wide_range:bool = True):

        self.vout_v = vout_v
        self.iout_a = iout_a
        self.vin_step = vin_step
        self.coupling = coupling
        self.num_step = num_step
        self.sheet_name = sheet_name
        self.wb_filepath = wb_filepath
        self.column_step = column_step

        self.col_vout = col_vout
        self.col_iout = col_iout
        self.cc_current_start_A = cc_current_start_A
        self.cc_voltage_start_V = cc_voltage_start_V
        self.cc_current_end_A = cc_current_end_A
        self.cc_voltage_end_V = cc_voltage_end_V
        
        self.x_axis_wide_range = x_axis_wide_range

    def generate(self):
        """Generates a CVCC plot on the data source workbook"""
        
        # Check the files , will raise Exception if not
        check_if_files_exist(
            template_sheet_name = self.template_sheet_name,
            datasrc_sheet_name = self.sheet_name,
            datasrc_wb_filepath = self.wb_filepath)
        
        # Load both the template workbook and data source workbook
        wbs = load_workbook(chart_template_path)
        wb = load_workbook(self.wb_filepath)

        # Get the sheet containing the chart template
        ws_template = wbs[self.template_sheet_name]
        
        # Reset chartsheet and get the worksheet object
        self.sheet_name = f"CVCC_{self.coupling}_{self.vout_v:g}V_{self.iout_a:g}A" 
        ws = reset_chartsheet(wb, )

        # Copy chart from template then close template workbook
        self.chart = deepcopy(ws_template._charts[0])
        wbs.close() 
        
        # Setup the chart axes, plot area and other settings
        self.setup_chart_settings()
        
        sheet_name = self.sheet_name
        num_step = self.num_step
        row_startdata = self.row_startdata

        # Add chart series per line voltage input
        for self.series_index, vin_freq in enumerate(self.vin_step):

            start_col = 2+self.column_step*self.series_index
            # start_col = get_column_letter(2+self.column_step*self.series_index)
            vin = vin_freq[0]
            # wb[self.sheet_name][f'{start_col}6'].value
            # series_title = f"{1} VAC"
            series_title = f'{vin:g} V{self.coupling.upper()}'

            xvalues = Reference(
                wb[sheet_name], 
                min_col=start_col + self.col_iout, 
                min_row=row_startdata, 
                max_row=row_startdata+num_step-1)
            values = Reference(
                wb[sheet_name], 
                min_col=start_col + self.col_vout, 
                min_row=row_startdata, 
                max_row=row_startdata+num_step-1)
            
            self.format_plotseries(series_title, xvalues, values)

        self.series_index += 1
        
        while len(self.chart.series) > self.series_index:
            self.chart.series.pop()
        
        # Add +-25 mA limits
        if self.cc_current_start_A is not None:
            
            upper_limit_start = round(self.cc_current_start_A + 0.025,6)
            upper_limit_end = round(self.cc_current_end_A + 0.025,6)
            lower_limit_start = round(self.cc_current_start_A - 0.025,6)
            lower_limit_end = round(self.cc_current_end_A - 0.025,6)
            ul_title = '+25 mA'
            ll_title = '-25 mA'
            
            series = deepcopy(self.chart.series[len(self.chart.series) -1])
            self.chart.series.append(series)
            values = f"={{{self.cc_voltage_start_V},{self.cc_voltage_end_V}}}"
            xvalues = f"={{{upper_limit_start},{upper_limit_end}}}"
            self.chart.series[len(self.chart.series)-1].xVal.numRef= NumRef()
            self.chart.series[len(self.chart.series)-1].xVal.numRef.f = xvalues
            self.chart.series[len(self.chart.series)-1].yVal.numRef = NumRef()
            self.chart.series[len(self.chart.series)-1].yVal.numRef.f = values
            self.chart.series[len(self.chart.series)-1].title.v = ul_title
            self.chart.series[len(self.chart.series)-1].graphicalProperties.line.solidFill = "92D050"
            self.chart.series[len(self.chart.series)-1].graphicalProperties.line.dashStyle = "dash"
            self.chart.series[len(self.chart.series)-1].marker = openpyxl.chart.marker.Marker(None)
            self.chart.series[len(self.chart.series)-1].order = self.series_index
            self.series_index += 1
            
            series = deepcopy(self.chart.series[len(self.chart.series) -1])
            self.chart.series.append(series)
            values = f"={{{self.cc_voltage_start_V},{self.cc_voltage_end_V}}}"
            xvalues = f"={{{lower_limit_start},{lower_limit_end}}}"
            self.chart.series[len(self.chart.series)-1].xVal.numRef= NumRef()
            self.chart.series[len(self.chart.series)-1].xVal.numRef.f = xvalues
            self.chart.series[len(self.chart.series)-1].yVal.numRef = NumRef()
            self.chart.series[len(self.chart.series)-1].yVal.numRef.f = values
            self.chart.series[len(self.chart.series)-1].title.v = ll_title
            self.chart.series[len(self.chart.series)-1].graphicalProperties.line.solidFill = "FF0000"
            self.chart.series[len(self.chart.series)-1].graphicalProperties.line.dashStyle = "dash"
            self.chart.series[len(self.chart.series)-1].marker = openpyxl.chart.marker.Marker(None)
            self.chart.series[len(self.chart.series)-1].order = self.series_index
            
            self.series_index += 1
        
        while len(self.chart.series) > self.series_index:
            self.chart.series.pop()
            
        #Clear chart title then place to chartsheet and save workbook    
        self.chart.title = " "    
        ws.add_chart(self.chart)
        wb.save(self.wb_filepath)
        # return wb
    
    def setup_chart_settings(self):
        # Initialize chart layout properties
        xLabel = "Load (A)"
        yLabel = "Output Voltage (V)"
        
        if  self.vout_v <= 4.5:
            y_min_scale = 0
            y_max_scale = 0.5*(math.ceil((self.vout_v + 0.5)/0.5))
            #y_max_scale = 5
            y_major_unit = 0.5
            y_numFmt = '0.0'
        elif (self.vout_v > 4.5) & (self.vout_v <= 9):
            y_min_scale = 0
            y_max_scale = 1*(math.ceil((self.vout_v + 1)/1))
            #y_max_scale = 10
            y_major_unit = 1
            y_numFmt = '0'
        elif (self.vout_v > 9) & (self.vout_v <=18):
            y_min_scale = 0
            y_max_scale = 2*(math.ceil((self.vout_v + 2)/2))
            #y_max_scale = 20
            y_major_unit = 2
            y_numFmt = '0'
        elif (self.vout_v > 18) & (self.vout_v <=27):
            y_min_scale = 0
            y_max_scale = 3*(math.ceil((self.vout_v + 3)/3))
            #y_max_scale = 30
            y_major_unit = 3
            y_numFmt = '0'
        else:
            y_min_scale = 0
            y_max_scale = 5*(math.ceil((self.vout_v + 5)/5))
            #y_max_scale = 50
            y_major_unit = 5
            y_numFmt = '0'

        if self.x_axis_wide_range:
  
            if self.iout_a > 3:
                x_major_unit = 0.5
            elif (self.iout_a > 1) & (self.iout_a <= 3):
                x_major_unit = 0.3
            else:
                x_major_unit = 0.1
            x_numFmt = '0.0'

            x_min_scale = 0
            x_max_scale = x_major_unit*(math.ceil(self.iout_a/x_major_unit)+3)
        
        else:
            x_major_unit = 0.025
            x_numFmt = '0.000'

            x_min_scale = self.iout_a - 0.1
            x_max_scale = self.iout_a + 0.1

        # Set up the chart settings
        self.chart = chart_setup(
            self.chart, " ", xLabel, yLabel,
            x_min_scale, x_max_scale, y_min_scale, y_max_scale,
            x_major_unit, 1, y_major_unit, 1, 
            x_numFmt, y_numFmt)

    def format_plotseries(self, series_title, xvalues, values):
        s_index = self.series_index

        if s_index <= 3:
            self.chart.series[s_index] = edit_series(
                self.chart.series[s_index],
                xvalues,values,
                series_title,marker=None)
        else:
            series = deepcopy(self.chart.series[s_index-1])

            series.graphicalProperties.line.solidFill = color_list[s_index-4]
            series = edit_series(series,xvalues,values,series_title)
            series.marker.spPr.solidFill = color_list[s_index-4]
            series.marker.spPr.ln = LineProperties(noFill=True)
            self.chart.series.append(series)

def generate_plots_CVCC(
    vout: float,
    iout: float,
    num_step: int,
    vin_step: int,
    # wb: Workbook,
    sheet_name: str,
    wb_filepath: str,
    column_step: int
    ):
    
    """
    vout            : output voltage setpoint, numerical input in V
    iout            : nominal output current, numerical input in A
    num_step        : number of load step per sheet, integer input
    vin_step        : number of line input step per sheet, integer input
    wb              : workbook containing the data, must be openpyxl.workbook.Workbook object
    sheet_name      : sheet within the workbook containing the data, string input
    wb_filepath     : path and filename of excel file where charts will be saved, string input
    """
    # dst = wb_filepath
    template_sheet_name = 'Charts_Template'
    
    # Check the files , will raise Exception if not
    check_if_files_exist(
        template_sheet_name = template_sheet_name,
        datasrc_sheet_name = sheet_name,
        datasrc_wb_filepath = wb_filepath)
    
    # Load both the template workbook and data source workbook
    wbs = load_workbook(chart_template_path)
    wb = load_workbook(wb_filepath)

    # Get the sheet containing the chart template
    ws_template = wbs[template_sheet_name]
    
    # Reset chartsheet and get the worksheet object
    ws = reset_chartsheet(wb, sheet_name)

    # Copy chart from template then close template workbook
    chart = deepcopy(ws_template._charts[0])
    wbs.close() 
    
    # Initialize chart layout properties
    xLabel = "Load (A)"
    yLabel = "Output Voltage (V)"
    
    if  vout <= 4.5:
        y_min_scale = 0
        y_max_scale = 0.5*(math.ceil((vout + 0.5)/0.5))
        #y_max_scale = 5
        y_major_unit = 0.5
        y_numFmt = '0.0'
    elif (vout > 4.5) & (vout <= 9):
        y_min_scale = 0
        y_max_scale = 1*(math.ceil((vout + 1)/1))
        #y_max_scale = 10
        y_major_unit = 1
        y_numFmt = '0'
    elif (vout > 9) & (vout <=18):
        y_min_scale = 0
        y_max_scale = 2*(math.ceil((vout + 2)/2))
        #y_max_scale = 20
        y_major_unit = 2
        y_numFmt = '0'
    elif (vout > 18) & (vout <=27):
        y_min_scale = 0
        y_max_scale = 3*(math.ceil((vout + 3)/3))
        #y_max_scale = 30
        y_major_unit = 3
        y_numFmt = '0'
    else:
        y_min_scale = 0
        y_max_scale = 5*(math.ceil((vout + 5)/5))
        #y_max_scale = 50
        y_major_unit = 5
        y_numFmt = '0'
    iout = iout/1000    
    if iout > 3:
        x_major_unit = 0.5
    elif (iout > 1) & (iout <= 3):
        x_major_unit = 0.3
    else:
        x_major_unit = 0.1
    x_numFmt = '0.0'

    x_min_scale = 0
    x_max_scale = x_major_unit*(math.ceil(iout/x_major_unit)+3)

    # Set up the chart settings
    chart = chart_setup(
        chart, " ", xLabel, yLabel,
        x_min_scale, x_max_scale, y_min_scale, y_max_scale,
        x_major_unit, 1, y_major_unit, 1, 
        x_numFmt, y_numFmt)
    
    # Add chart series per line voltage input
    for series_index in range(vin_step):
        wb[sheet_name][f'{get_column_letter(2+14*series_index)}6'].value
        series_title = f"{1} VAC"
        xvalues = Reference(wb[sheet_name], min_col=10+14*series_index, min_row=6, max_row=6+num_step-1)
        values = Reference(wb[sheet_name], min_col=9+14*series_index, min_row=6, max_row=6+num_step-1)
        if series_index <= 3:
            chart.series[series_index] = edit_series(chart.series[series_index],xvalues,values,series_title,marker=None)
        else:
            series = deepcopy(chart.series[series_index-1])
            chart.series.append(series)
            chart.series[series_index].graphicalProperties.line.solidFill = color_list[series_index-4]
            chart.series[series_index] = edit_series(chart.series[series_index],xvalues,values,series_title)
            chart.series[series_index].marker.spPr.solidFill = color_list[series_index-4]
            chart.series[series_index].marker.spPr.ln = LineProperties(noFill=True)
    series_index = series_index + 1
    while len(chart.series) > series_index:
        chart.series.pop()

    #Clear chart title then place to chartsheet and save workbook    
        
    chart.title = " "    
    ws.add_chart(chart)

    wb.save(wb_filepath)
    return wb

def generate_plots_InputHarmonics(
    vout: float,
    iout: float,
    vin_step: int,
    wb: Workbook,
    sheet_name: str,
    wb_filepath: str):
    
    """
    vout            : output voltage setpoint, numerical input in V
    iout            : nominal output current, numerical input in A
    vin_step        : number of line input step per sheet, integer input
    wb              : workbook containing the data, must be openpyxl.workbook.Workbook object
    sheet_name      : sheet within the workbook containing the data, string input
    wb_filepath     : path and filename of excel file where charts will be saved, string input
    """        
    
    template_sheet_name = 'Charts_Template_Input_Harmonics'
    
    # Check the files , will raise Exception if not
    check_if_files_exist(
        template_sheet_name = template_sheet_name,
        datasrc_sheet_name = sheet_name,
        datasrc_wb_filepath = wb_filepath)
    
    # Load both the template workbook and data source workbook
    wbs = load_workbook(chart_template_path)
    wb = load_workbook(wb_filepath)

    # Get the sheet containing the chart template
    ws_template = wbs[template_sheet_name]

    #Reset chartsheet
    for Vin_index in range (vin_step):
        Vin = wb[sheet_name][f'B{6+30*Vin_index}'].value
        chart_sheet_name = f'In_Harm_{vout:g}V_{iout:g}A_{Vin:g}_VAC'
        ws = reset_chartsheet(wb, chart_sheet_name)

        #Copy chart from template then close template workbook

        chart = deepcopy(ws_template._charts[0])
        # chart.plot_area.spPr = GraphicalProperties()
        # chart.plot_area.spPr.ln.solidFill = ColorChoice(srgbClr = "000000")
        # chart.plot_area.spPr.ln.w=50761 #4 pt border thickness
        
        chart.x_axis.title.text.rich.p[0].r[0].rPr = CharacterProperties(b = True,latin = Font(typeface= 'Calibri'), sz=1400)
        cp = CharacterProperties(sz=1200, latin=Font(typeface='Calibri'),b=True)
        chart.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
        chart.y_axis.title.text.rich.p[0].r[0].rPr = CharacterProperties(b= True,latin = Font(typeface= 'Calibri'), sz=1400)
        chart.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
        
        # Assign the RichText properties to the chart legend
        chart.legend.textProperties = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
        
        values = Reference(wb[sheet_name], min_col=3, min_row=12+Vin_index*30, max_row=30+Vin_index*30)
        chart.series[0] = SeriesFactory(values=values,title = 'Harmonic content (mA)' )
        chart.series[0].val.numRef.f = values
        chart.series[0].cat= AxDataSource(strRef=StrRef(f = '={3,5,7,9,11,13,15,17,19,21,23,25,27,29,31,33,35,37,39}')) 
        chart.series[0].spPr = deepcopy(ws_template._charts[0].series[0].spPr)

        values = Reference(wb[sheet_name], min_col=5, min_row=12+Vin_index*30, max_row=30+Vin_index*30)
        chart.series[1]= SeriesFactory(values=values,title  = 'Class D Limit (mA)' )
        chart.series[1].val.numRef.f  = values
        chart.series[1].spPr = deepcopy(ws_template._charts[0].series[1].spPr)
        
        #chart.set_categories('={3,5,7,9,11,13,15,17,19,21,23,25,27,29,31,33,35,37,39}')
        
        #chart.title = " "  
        ws.add_chart(chart)   
 
    wb.save(wb_filepath)    
    wbs.close()
    return wb

def generate_plots_NoLoad(
    vin_step: int,
    wb: Workbook,
    sheet_name: str,
    coupling: str,
    wb_filepath: str):
    
    """
    vin_step        : number of line input step per sheet, integer input
    wb              : workbook containing the data, must be openpyxl.workbook.Workbook object
    sheet_name      : sheet within the workbook containing the data, string input
    coupling        : coupling of input voltage, either "DC" or "AC"
    wb_filepath     : path and filename of excel file where charts will be saved, string input
    """        

    template_sheet_name = 'Charts_Template_No_Load'
    
    # Check the files , will raise Exception if not
    check_if_files_exist(
        template_sheet_name = template_sheet_name,
        datasrc_sheet_name = sheet_name,
        datasrc_wb_filepath = wb_filepath)
    
    # Load both the template workbook and data source workbook
    wbs = load_workbook(chart_template_path)
    wb = load_workbook(wb_filepath)

    # Get the sheet containing the chart template
    ws_template = wbs[template_sheet_name]
    
    # Reset chartsheet and get the worksheet object
    ws = reset_chartsheet(wb, sheet_name)

    # Copy chart from template then close template workbook
    chart = deepcopy(ws_template._charts[0])
    wbs.close() 

    xLabel = f"Input Voltage (V{coupling.upper()})"
    # chart.x_axis.title.text.rich.p[0].r[0].t = xLabel
    # chart.x_axis.title.text.rich.p[0].r[0].rPr = CharacterProperties(b = True,latin = Font(typeface= 'Calibri'), sz=1400)
    
    data_no_load = excel_to_df(wb_filepath,sheet_name,'F6',f'F{6+(vin_step-1)}')
    y_max_scale = (math.ceil(max(data_no_load.values)/20)+1)*20
    if y_max_scale < 100:
        y_max_scale = 100
    y_major_unit = y_max_scale/10
    
    chart.y_axis.scaling.max = y_max_scale
    chart.y_axis.majorUnit = y_major_unit
    
    # chart.plot_area.spPr = GraphicalProperties()
    # chart.plot_area.spPr.ln.solidFill = ColorChoice(srgbClr = "000000")
    # chart.plot_area.spPr.ln.w=50761 #4 pt border thickness
    
    chart.x_axis.title.text.rich.p[0].r[0].t = xLabel
    chart.x_axis.title.text.rich.p[0].r[0].rPr = CharacterProperties(b = True,latin = Font(typeface= 'Calibri'), sz=1400)
    cp = CharacterProperties(sz=1200, latin=Font(typeface='Calibri'),b=True)
    chart.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    chart.y_axis.title.text.rich.p[0].r[0].rPr = CharacterProperties(b= True,latin = Font(typeface= 'Calibri'), sz=1400)
    chart.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    
    values = Reference(wb[sheet_name], min_col=6, min_row=6, max_row=6+vin_step-1)
    chart.ser[0] = SeriesFactory(values=values)
    chart.ser[0].val.numRef.f = values
    chart.ser[0].cat= AxDataSource(strRef=StrRef(f = Reference(wb[sheet_name], min_col=2, min_row=6, max_row=6+vin_step-1)))
    chart.ser[0].spPr = deepcopy(ws_template._charts[0].ser[0].spPr)
    chart.ser[0].labels = deepcopy(ws_template._charts[0].ser[0].labels)    
    chart.title = " "  
    ws.add_chart(chart)   
    
    wb.save(wb_filepath)    
    return wb

def generate_plots_LineSense(
    vout: float,
    iout: float,
    num_step: int,
    vin_step: int,
    max_vin: float,
    min_vin: float,
    coupling: str,
    wb: Workbook,
    sheet_name: str,
    wb_filepath: str):
    
    """
    vout            : output voltage setpoint, numerical input
    iout            : output current setpoint, numerical input
    num_step        : number of load step per sheet, integer input
    vin_step        : number of line input step per sheet, integer input
    max_vin         : maximum input voltage, float input
    min_vin         : minimum input voltage, float input
    coupling        : coupling of input voltage, either "DC" or "AC"
    wb              : workbook containing the data, must be openpyxl.workbook.Workbook object
    sheet_name      : sheet within the workbook containing the data, string input
    wb_filepath     : path and filename of excel file where charts will be saved, string input
    """
    template_sheet_name = 'Charts_Template_Eff_v_Line'
    
    # Check the files , will raise Exception if not
    check_if_files_exist(
        template_sheet_name = template_sheet_name,
        datasrc_sheet_name = sheet_name,
        datasrc_wb_filepath = wb_filepath)
    
    # Load both the template workbook and data source workbook
    wbs = load_workbook(chart_template_path)
    wb = load_workbook(wb_filepath)

    # Get the sheet containing the chart template
    ws_template = wbs[template_sheet_name]
    
    # Reset chartsheet and get the worksheet object
    chart_sheet_name = f"LineSense_{coupling}_{vout:g}V_{iout:g}A" 
    ws = reset_chartsheet(wb, chart_sheet_name)

    # Copy chart from template then close template workbook
    chart = deepcopy(ws_template._charts[0])
    wbs.close() 
    
    #Initialize chart layout properties
    
    xLabel = f"Input Voltage (V{coupling.upper()})"
    yLabel = "Calculated Input Voltage (VDC)"
    
    y_min_scale = 0
    y_major_unit = 5*math.ceil(max_vin/100)
    y_max_scale = y_major_unit*(math.ceil(max_vin/y_major_unit)+1)
    y_numFmt = '0'
    x_major_unit = 5*math.ceil(max_vin/100)
    x_min_scale = min_vin - x_major_unit
    x_max_scale = min_vin + (math.ceil((max_vin-x_min_scale)/x_major_unit)+1)*x_major_unit
    x_numFmt = '0'
    
    chart = chart_setup(chart," ",xLabel,yLabel,
                x_min_scale,x_max_scale,y_min_scale,y_max_scale,x_major_unit,
                1,y_major_unit,1,x_numFmt,y_numFmt)
    
    #Add chart series per line voltage input
        
    for series_index in range(num_step):
        series_title = f"{wb[sheet_name][f'A{6+(vin_step+4)*series_index}'].value}% Load"
        xvalues = Reference(wb[sheet_name], min_col=4, min_row=6+((vin_step+4)*series_index), max_row=(6+(vin_step-1)+((vin_step+4)*series_index)))
        values = Reference(wb[sheet_name], min_col=13, min_row=6+((vin_step+4)*series_index), max_row=(6+(vin_step-1)+((vin_step+4)*series_index)))
        if series_index <= 3:
            chart.series[series_index] = edit_series(chart.series[series_index],xvalues,values,series_title)
        else:
            series = deepcopy(chart.series[series_index-1])
            chart.series.append(series)
            chart.series[series_index].graphicalProperties.line.solidFill = color_list[series_index-4]
            chart.series[series_index] = edit_series(chart.series[series_index],xvalues,values,series_title)
            chart.series[series_index].marker.spPr.solidFill = color_list[series_index-4]
            chart.series[series_index].marker.spPr.ln = LineProperties(noFill=True)
    series_index = series_index + 1
    
    #Remove extra chart series (if vin_step < 4)
    
    while len(chart.series) > (series_index):
        chart.series.pop()
    
    #Clear chart title then place to chartsheet and save workbook
    
    chart.title = " "  
    ws.add_chart(chart)        
    wb.save(wb_filepath)
    return wb

def generate_table_AveEff(
    vin_step: int,
    coupling: str,
    sheet_name: str,
    wb_filepath: str):
    
    """
    vin_step            : number of line input step per sheet, integer input
    coupling            : coupling of input voltage, either "DC" or "AC"
    sheet_name          : sheet within the workbook containing the data, string input
    wb_filepath         : path and filename of excel file where charts will be saved, string input
    """        

    #Check if files and folders exist
    src = wb_filepath
    dst = wb_filepath
    
    wb = load_workbook(src)
    if 'Template_Eff_Table' not in wb.sheetnames:
        print('Template_Eff_Table does not exist')
        return
    ws_template = wb['Template_Eff_Table']
    
    if sheet_name not in wb.sheetnames:
        print(f'sheet "{sheet_name}" not in workbook')
        return
    if f"Table_{sheet_name}" in wb.sheetnames:
        wb.remove(wb[f"Table_{sheet_name}"])
    ws = wb.copy_worksheet(ws_template)
    ws.title = f"Table_{sheet_name}"
    wb.move_sheet(f"Table_{sheet_name}",-(len(wb.sheetnames)-1)+1)
    df = excel_to_df(src, sheet_name, "A6", f"M{6+vin_step*6-2}")
    df2 = excel_to_df(src, sheet_name, "N6", f"R{6+vin_step*6-2}")
    
    vin_list = df.loc[:,[f"V{coupling.upper()} (rms)"]]
    df = df.loc[:,['Load','Pin (W)','Vo (V)','Io (A)','Po (W)','Efficiency']]
    
    df_row_len, df_col_len = df.shape
    end_row = df_row_len - 1
    end_col = df_col_len - 1

    for row in range(0, end_row+1):
        for col in range(0, end_col+1):
            if df.columns[col] not in ['Load','VAC (rms)','VDC (rms)','Freq (Hz)','V Reg(5%)']:
                try:
                    temp = f"{(df.iloc[row, col]):#.3g}"
                    if temp[-1] == '.':
                        df.iloc[row, col] = temp[:-1]
                    else:
                        df.iloc[row, col] = temp
                except:
                    pass

    df_row_len, df_col_len = df2.shape
    end_row = df_row_len - 1
    end_col = df_col_len - 1

    for row in range(0, end_row+1):
        for col in range(0, end_col+1):
            if df2.columns[col] not in ['Load','VAC (rms)','VDC (rms)','Freq (Hz)','V Reg(5%)']:
                try:
                    temp = f"{(df2.iloc[row, col]):#.3g}"
                    if temp[-1] == '.':
                        df2.iloc[row, col] = temp[:-1]
                    else:
                        df2.iloc[row, col] = temp
                except:
                    pass
    
    for Vin_index in range(vin_step):
        eff_table = df.loc[0+Vin_index*6:4+Vin_index*6,:]
        eff_table_ave = df2.loc[3+Vin_index*6:3+Vin_index*6,['Avg_Eff','DOE6 Limit', 'COC5 T2 Limit','Pass/Fail']]
        eff_table_ave['Pass/Fail'].values[0] = (eff_table_ave['Pass/Fail'].values[0])[0:4]
        eff_table_10pct = df2.loc[4+Vin_index*6:4+Vin_index*6,['COC5_T2_10%','Pass/Fail']]
        eff_table_10pct['Pass/Fail'].values[0] = (eff_table_10pct['Pass/Fail'].values[0])[0:4]
        wb[f"Table_{sheet_name}"][f'B{6+Vin_index*5}'] = vin_list.loc[0+Vin_index*6].values[0]
        df_to_excel(wb, f"Table_{sheet_name}", eff_table, f'C{6+Vin_index*5}')
        df_to_excel(wb, f"Table_{sheet_name}", eff_table_ave, f'I{6+Vin_index*5}')
        df_to_excel(wb, f"Table_{sheet_name}", eff_table_10pct, f'K{10+Vin_index*5}')
    # change_number_format(wb, f"Table_{sheet_name}",6, 5, 25, 6,"#,##0.000")
    # change_number_format(wb, f"Table_{sheet_name}",6, 9, 25, 11,"#,##0.00")
    wb.remove(wb['Template_Eff_Table'])
    wb.save(dst)   
    
    wb.close()

def chart_setup(chart: ChartBase, title: str=" ", xLabel: str=" ", yLabel: str=" ",
                x_min_scale: float=0, x_max_scale: float=100, 
                y_min_scale: float=0,y_max_scale: float=100,
                x_major_unit: float=1, x_minor_unit: float=0.1,
                y_major_unit: float=1, y_minor_unit: float=0.1,
                x_numFmt: str='0', y_numFmt: str='0',
                height: float=6.65*2.541, width: float=10.18*2.541):
    
    #chart is the chart object to be setup
    #title is the chart title, string input
    #xLabel and yLabel correspond to the labels of the x and y-axis, respectively, string input
    #x_min_scale and x_max_scale correspond to mininum and maximum values, respectively, in the x-axis, numerical input
    #y_min_scale and y_max_scale correspond to mininum and maximum values, respectively, in the y-axis, numerical input
    #y_major_unit and y_minor_unit correspond to the major and minor interval length per division in the x-axis, numerical input
    #y_major_unit and y_minor_unit correspond to the major and minor interval length per division in the y-axis,numerical input
    #height and weight is the dimensions of the chart area in cm, numerical input
    #x_numFmt and y_numFmt is the number format of the for the axis (eg. '0.000', '0.0', etc.), string input
    
    # Format chart title, axis, and axis labels 
    chart.title = title
    chart.x_axis.title.text.rich.p[0].r[0].t = xLabel
    chart.x_axis.title.text.rich.p[0].r[0].rPr = CharacterProperties(b = True,latin = Font(typeface= 'Calibri'), sz=1400)
    cp = CharacterProperties(sz=1200, latin=Font(typeface='Calibri'),b=True)
    chart.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    chart.y_axis.title.text.rich.p[0].r[0].t = yLabel
    chart.y_axis.title.text.rich.p[0].r[0].rPr = CharacterProperties(b= True,latin = Font(typeface= 'Calibri'), sz=1400)
    chart.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    
    # Assign the RichText properties to the chart legend
    chart.legend.textProperties = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])

    chart.x_axis.scaling.min = x_min_scale
    chart.x_axis.scaling.max = x_max_scale
    chart.y_axis.scaling.min = y_min_scale
    chart.y_axis.scaling.max = y_max_scale
    
    chart.x_axis.majorUnit = x_major_unit
    chart.x_axis.miorUnit = x_minor_unit
    chart.y_axis.majorUnit = y_major_unit
    chart.x_axis.minorUnit = y_minor_unit
    chart.x_axis.numFmt = x_numFmt
    chart.y_axis.numFmt = y_numFmt

    # chart.plot_area.spPr = GraphicalProperties()

    # chart.plot_area.spPr.bwMode = 'black'
    # chart.plot_area.spPr.ln.solidFill = ColorChoice(srgbClr = "000000")
    chart.height = height
    chart.width = width
    # chart.plot_area.spPr.ln.w=50761 #4 pt border thickness

    # chart.plot_area.spPr.ln.cmpd ='sng'
    # chart.plot_area.spPr.ln.cap ='flat'

    return chart
    
def edit_series(series: Series=None,xVal: Reference=None,yVal: Reference=None,title: str='',marker: str='auto'):
    
    #series is the chart series to be edited see (see openpyxl.chart.Series)
    #xVal is the Reference for the x-coordinates (see openpyxl.chart.Reference)
    #yVal is the Reference for the y-coordinates (see openpyxl.chart.Reference)
    #title is the series title in str format
    #marker is the marker for the datapoints in the chart (eg.‘triangle’, ‘dash’, ‘dot’, ‘star’, ‘circle’, ‘picture’, ‘square’, ‘x’, ‘plus’, ‘auto’, ‘diamond’)
    
    temp = SeriesFactory(values=yVal, xvalues=xVal)
    series.yVal = temp.yVal
    series.xVal = temp.xVal
    series.title.v = title
    if marker is None:
            series.marker = openpyxl.chart.marker.Marker(None)
    else: series.marker.symbol = marker
    return series